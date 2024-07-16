import requests
import os.path
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font

now = datetime.now()
nowMinusOneDay = datetime.now() - timedelta(days=1)
actualMonth = now.strftime("%m")
lastMonth = nowMinusOneDay.strftime("%m")
year = nowMinusOneDay.strftime("%Y")
pathFile = "/media/Statistics_ESA"


class AsyncOsApiJsonFile:
    def __init__(self):
        self.username = "<SMA API username>"
        self.password = "<SMA API password>"
        self.url = None

    def set_json(self, resource, last_month, actual_month):
        url = f"https://<SMA_APPLIANCE_FQDN>:6443/sma/api/v2.0/reporting/{resource}?startDate=" \
              f"2021-{last_month}-01T00:00:00.000Z&endDate=2021-{actual_month}-01T00:00:00.000Z&device_type=esa"
        r = requests.get(url, auth=(self.username, self.password), verify=False)
        return r.json()


if __name__ == "__main__":
    monthlyStats = AsyncOsApiJsonFile()

    json_in = monthlyStats.set_json("mail_incoming_traffic_summary", lastMonth, actualMonth)
    json_out = monthlyStats.set_json("mail_outgoing_traffic_summary", lastMonth, actualMonth)

    if os.path.isfile('{}/Statistics_ESA_{}.xlsx'.format(pathFile, year)):
        workbook = load_workbook('{}/Statistics_ESA_{}.xlsx'.format(pathFile, year))
    else:
        workbook = Workbook()
        ws = workbook.active
        for i in range(1, 13):
            ws = workbook.create_sheet(str(i).zfill(2))
            ws['A1'] = "Incoming Total threat recipients"
            ws['B1'] = "Incoming Total safe recipients"
            ws['C1'] = "Outgoing Total delivered recipients"
            for cell in ws["1:1"]:
                cell.font = Font(bold=True)
        workbook.remove(workbook['Sheet'])

    ws = workbook[lastMonth]
    ws['A2'] = total_threat_recipients = json_in["data"]["resultSet"][15]["total_threat_recipients"]
    ws['B2'] = json_in["data"]["resultSet"][16]["total_recipients"] - total_threat_recipients
    ws['C2'] = json_out["data"]["resultSet"][14]["total_recipients_delivered"]
    workbook.save('{}/Statistics_ESA_{}.xlsx'.format(pathFile, year))
